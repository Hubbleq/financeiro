import tempfile
import os
import sys
import pandas as pd
from openpyxl import load_workbook

# Garantir que o diretório do projeto esteja no path para importar o módulo
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from analise_financeira import AnaliseFinanceira


def test_export_no_duplicated_header():
    analise = AnaliseFinanceira()
    # criar um DataFrame mínimo que satisfaça gerar_relatorio
    df_result = pd.DataFrame([
        {
            'Nome_Produto': 'Teste',
            'Custo_Compra_Unitario': 10.0,
            'Quantidade': 2,
            'custo_total_compra': 20.0,
            'custo_fixo_total': 0.0,
            'custo_total_geral': 20.0,
            'preco_venda_total': 25.0,
            'lucro_total': 5.0,
            'markup_porcentagem': 25.0,
            'margem_lucro_real': 20.0
        }
    ])

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    caminho = tmp.name

    try:
        analise.exportar_resultados(df_result, caminho)

        wb = load_workbook(caminho, read_only=True)
        ws = wb['Analise_Produtos']

        # linha 1: título (pode estar mesclada)
        # linha 2: cabeçalho (escrito pelo pandas)
        # linha 3: primeiro registro de dados
        header_row = [cell.value for cell in ws[2]]
        first_data_row = [cell.value for cell in ws[3]]

        # verificar que o cabeçalho e a primeira linha de dados não são idênticos
        assert header_row != first_data_row, "Cabeçalho duplicado detectado: header_row == first_data_row"

    finally:
        try:
            os.remove(caminho)
        except Exception:
            pass
